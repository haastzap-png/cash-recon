from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .logic import CashReconResult


def _fmt_dt(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _fmt_d(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d")


def _money(ws, cell: str):
    ws[cell].number_format = "#,##0"


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col[:5000]:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)


def build_cash_recon_workbook(result: CashReconResult) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    title_font = Font(size=14, bold=True)
    header_font = Font(bold=True)
    bad_fill = PatternFill("solid", fgColor="FCE4D6")
    ok_fill = PatternFill("solid", fgColor="E2EFDA")

    ws["A1"] = "現金對帳表"
    ws["A1"].font = title_font

    ws["A3"] = "分店"
    ws["B3"] = result.store
    ws["A4"] = "區間開始"
    ws["B4"] = _fmt_dt(result.period.start)
    ws["A5"] = "區間結束"
    ws["B5"] = _fmt_dt(result.period.end)

    ws["A7"] = "漏結帳筆數(已報到但帳單金額空)"
    ws["B7"] = len(result.missing_bills)
    ws["A8"] = "是否可視為正確現金對帳"
    ws["B8"] = "是" if len(result.missing_bills) == 0 else "否"
    ws["B8"].fill = ok_fill if len(result.missing_bills) == 0 else bad_fill

    ws["A10"] = "Hotcake 服務現金(依訂單日期時間區間)"
    ws["B10"] = result.totals.hotcake_service_cash
    _money(ws, "B10")
    ws["A11"] = "Hotcake 儲值金現金(依結帳操作時間區間)"
    ws["B11"] = result.totals.hotcake_topup_cash
    _money(ws, "B11")
    ws["A12"] = "Hotcake 現金合計"
    ws["B12"] = result.totals.hotcake_cash_total
    _money(ws, "B12")

    ws["A14"] = "收銀機現金合計(依建立時間區間)"
    ws["B14"] = result.totals.pos_cash_total if result.totals.pos_cash_total is not None else ""
    if result.totals.pos_cash_total is not None:
        _money(ws, "B14")
    ws["A15"] = "收銀機現金 - Hotcake 現金合計"
    ws["B15"] = result.totals.pos_cash_diff if result.totals.pos_cash_diff is not None else ""
    if result.totals.pos_cash_diff is not None:
        _money(ws, "B15")
    ws["A16"] = "時間容忍外(Hotcake)筆數"
    ws["B16"] = len(result.hotcake_time_mismatches)
    ws["A17"] = "時間容忍外(POS)筆數"
    ws["B17"] = len(result.pos_time_mismatches)
    ws["A18"] = "刷卡機實付合計(依交易時間區間)"
    ws["B18"] = result.totals.card_machine_total if result.totals.card_machine_total is not None else ""
    if result.totals.card_machine_total is not None:
        _money(ws, "B18")
    ws["A19"] = "刷卡機對帳未匹配(卡機)"
    ws["B19"] = len([r for r in result.card_mismatches if r.source == "card"])
    ws["A20"] = "刷卡機對帳未匹配(Hotcake)"
    ws["B20"] = len([r for r in result.card_mismatches if r.source == "hotcake"])

    for cell in (
        "A3",
        "A4",
        "A5",
        "A7",
        "A8",
        "A10",
        "A11",
        "A12",
        "A14",
        "A15",
        "A16",
        "A17",
        "A18",
        "A19",
        "A20",
    ):
        ws[cell].font = header_font
    ws["A1"].alignment = Alignment(horizontal="left")
    _auto_width(ws)

    ws_missing = wb.create_sheet("MissingBills")
    ws_missing.append(
        [
            "分店",
            "日期(服務開始)",
            "訂單編號(數字)",
            "訂單編號(代碼)",
            "日期時間(服務開始)",
            "設計師",
            "服務",
            "訂單狀態",
            "報到/取消時間",
            "會員姓名",
            "手機號碼",
        ]
    )
    for c in range(1, 12):
        ws_missing.cell(1, c).font = header_font
    for r in result.missing_bills:
        ws_missing.append(
            [
                r.store,
                r.service_start.strftime("%Y-%m-%d"),
                r.order_id,
                r.order_code,
                _fmt_dt(r.service_start),
                r.designer,
                r.service,
                r.order_status,
                _fmt_dt(r.checkin_time),
                r.member_name,
                r.phone,
            ]
        )
    _auto_width(ws_missing)

    ws_service = wb.create_sheet("HotcakeBills_Service")
    ws_service.append(
        ["分店", "帳單編號", "結帳操作時間", "計算歸屬日", "設計師", "項目", "現金", "信用卡", "Line Pay", "結帳金額"]
    )
    for c in range(1, 11):
        ws_service.cell(1, c).font = header_font
    for r in result.service_bill_rows:
        ws_service.append(
            [
                r.store,
                r.bill_id,
                _fmt_dt(r.settlement_time),
                _fmt_d(r.attributed_date),
                r.designer,
                r.item,
                r.cash,
                r.credit_card,
                r.linepay,
                r.bill_amount,
            ]
        )
    for row in ws_service.iter_rows(min_row=2, min_col=7, max_col=10):
        for cell in row:
            cell.number_format = "#,##0"
    _auto_width(ws_service)

    ws_topup = wb.create_sheet("HotcakeBills_Topup")
    ws_topup.append(
        ["分店", "帳單編號", "結帳操作時間", "計算歸屬日", "設計師", "項目", "現金", "信用卡", "Line Pay", "結帳金額"]
    )
    for c in range(1, 11):
        ws_topup.cell(1, c).font = header_font
    for r in result.topup_bill_rows:
        ws_topup.append(
            [
                r.store,
                r.bill_id,
                _fmt_dt(r.settlement_time),
                _fmt_d(r.attributed_date),
                r.designer,
                r.item,
                r.cash,
                r.credit_card,
                r.linepay,
                r.bill_amount,
            ]
        )
    for row in ws_topup.iter_rows(min_row=2, min_col=7, max_col=10):
        for cell in row:
            cell.number_format = "#,##0"
    _auto_width(ws_topup)

    ws_hm = wb.create_sheet("TimeMismatch_Hotcake")
    ws_hm.append(
        [
            "分店",
            "日期(服務開始)",
            "日期時間(服務開始)",
            "設計師",
            "服務",
            "分鐘",
            "帳單編號",
            "帳單金額",
            "現金",
            "現金差額",
            "最近POS時間",
            "時間差(分鐘)",
            "原因",
        ]
    )
    for c in range(1, 14):
        ws_hm.cell(1, c).font = header_font
    for r in result.hotcake_time_mismatches:
        ws_hm.append(
            [
                r.store,
                r.service_start.strftime("%Y-%m-%d"),
                _fmt_dt(r.service_start),
                r.designer,
                r.service,
                r.minutes if r.minutes is not None else "",
                r.bill_id,
                r.bill_amount,
                r.cash,
                r.cash_diff if r.cash_diff is not None else "",
                _fmt_dt(r.nearest_pos_time),
                r.nearest_diff_minutes if r.nearest_diff_minutes is not None else "",
                r.reason,
            ]
        )
    for row in ws_hm.iter_rows(min_row=2, min_col=8, max_col=10):
        for cell in row:
            cell.number_format = "#,##0"
    _auto_width(ws_hm)

    ws_pm = wb.create_sheet("TimeMismatch_POS")
    ws_pm.append(
        [
            "機台名稱",
            "日期(建立時間)",
            "日期時間(建立時間)",
            "設計師",
            "商品名稱",
            "分鐘",
            "現金支付",
            "現金差額",
            "最近Hotcake時間",
            "時間差(分鐘)",
            "原因",
        ]
    )
    for c in range(1, 12):
        ws_pm.cell(1, c).font = header_font
    for r in result.pos_time_mismatches:
        ws_pm.append(
            [
                r.terminal_name,
                r.created_time.strftime("%Y-%m-%d"),
                _fmt_dt(r.created_time),
                r.designer,
                r.product_name,
                r.minutes if r.minutes is not None else "",
                r.cash_paid,
                r.cash_diff if r.cash_diff is not None else "",
                _fmt_dt(r.nearest_hotcake_time),
                r.nearest_diff_minutes if r.nearest_diff_minutes is not None else "",
                r.reason,
            ]
        )
    for row in ws_pm.iter_rows(min_row=2, min_col=7, max_col=8):
        for cell in row:
            cell.number_format = "#,##0"
    _auto_width(ws_pm)

    if result.card_machine_rows:
        ws_card = wb.create_sheet("CardMachine_Transactions")
        ws_card.append(
            ["訂單編號", "店鋪名稱", "設備名稱", "交易金額", "實付金額", "交易時間", "支付方式"]
        )
        for c in range(1, 8):
            ws_card.cell(1, c).font = header_font
        for r in result.card_machine_rows:
            ws_card.append(
                [
                    r.order_id,
                    r.store,
                    r.device_name,
                    r.amount,
                    r.paid_amount,
                    _fmt_dt(r.transaction_time),
                    r.pay_method,
                ]
            )
        for row in ws_card.iter_rows(min_row=2, min_col=4, max_col=5):
            for cell in row:
                cell.number_format = "#,##0"
        _auto_width(ws_card)

    if result.card_matches:
        ws_card_match = wb.create_sheet("CardMachine_Match")
        ws_card_match.append(
            ["分店", "帳單編號", "支付方式", "Hotcake金額", "Hotcake時間", "刷卡機金額", "刷卡機時間", "時間差(分鐘)"]
        )
        for c in range(1, 9):
            ws_card_match.cell(1, c).font = header_font
        for r in result.card_matches:
            ws_card_match.append(
                [
                    r.store,
                    r.bill_id,
                    r.pay_type,
                    r.hotcake_amount,
                    _fmt_dt(r.hotcake_time),
                    r.card_amount,
                    _fmt_dt(r.card_time),
                    r.time_diff_minutes,
                ]
            )
        for row in ws_card_match.iter_rows(min_row=2, min_col=4, max_col=6):
            for cell in row:
                cell.number_format = "#,##0"
        _auto_width(ws_card_match)

    if result.card_mismatches:
        ws_card_mis = wb.create_sheet("CardMachine_Mismatch")
        ws_card_mis.append(
            ["來源", "分店", "支付方式", "金額", "時間", "帳單編號", "最近時間", "時間差(分鐘)"]
        )
        for c in range(1, 9):
            ws_card_mis.cell(1, c).font = header_font
        for r in result.card_mismatches:
            ws_card_mis.append(
                [
                    r.source,
                    r.store,
                    r.pay_type,
                    r.amount,
                    _fmt_dt(r.time),
                    r.bill_id,
                    _fmt_dt(r.nearest_time),
                    r.nearest_diff_minutes if r.nearest_diff_minutes is not None else "",
                ]
            )
        for row in ws_card_mis.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = "#,##0"
        _auto_width(ws_card_mis)

    # Summary by date and designer
    day_summary = {}
    for r in result.hotcake_time_mismatches:
        day = r.service_start.strftime("%Y-%m-%d")
        day_summary.setdefault(day, {"hotcake_count": 0, "pos_count": 0, "hotcake_cash": 0.0, "pos_cash": 0.0})
        day_summary[day]["hotcake_count"] += 1
        day_summary[day]["hotcake_cash"] += r.cash
    for r in result.pos_time_mismatches:
        day = r.created_time.strftime("%Y-%m-%d")
        day_summary.setdefault(day, {"hotcake_count": 0, "pos_count": 0, "hotcake_cash": 0.0, "pos_cash": 0.0})
        day_summary[day]["pos_count"] += 1
        day_summary[day]["pos_cash"] += r.cash_paid

    ws_day = wb.create_sheet("MismatchSummary_Day")
    ws_day.append(["日期", "Hotcake未匹配", "POS未匹配", "Hotcake未匹配現金", "POS未匹配現金"])
    for c in range(1, 6):
        ws_day.cell(1, c).font = header_font
    for day in sorted(day_summary.keys()):
        s = day_summary[day]
        ws_day.append([day, s["hotcake_count"], s["pos_count"], s["hotcake_cash"], s["pos_cash"]])
    for row in ws_day.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = "#,##0"
    _auto_width(ws_day)

    designer_summary = {}
    for r in result.hotcake_time_mismatches:
        key = r.designer or "(空白)"
        designer_summary.setdefault(key, {"hotcake_count": 0, "pos_count": 0, "hotcake_cash": 0.0, "pos_cash": 0.0})
        designer_summary[key]["hotcake_count"] += 1
        designer_summary[key]["hotcake_cash"] += r.cash
    for r in result.pos_time_mismatches:
        key = r.designer or "(空白)"
        designer_summary.setdefault(key, {"hotcake_count": 0, "pos_count": 0, "hotcake_cash": 0.0, "pos_cash": 0.0})
        designer_summary[key]["pos_count"] += 1
        designer_summary[key]["pos_cash"] += r.cash_paid

    ws_des = wb.create_sheet("MismatchSummary_Designer")
    ws_des.append(["設計師", "Hotcake未匹配", "POS未匹配", "Hotcake未匹配現金", "POS未匹配現金"])
    for c in range(1, 6):
        ws_des.cell(1, c).font = header_font
    for name in sorted(designer_summary.keys()):
        s = designer_summary[name]
        ws_des.append([name, s["hotcake_count"], s["pos_count"], s["hotcake_cash"], s["pos_cash"]])
    for row in ws_des.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = "#,##0"
    _auto_width(ws_des)

    return wb


def save_cash_recon_report(result: CashReconResult, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_cash_recon_workbook(result)
    wb.save(out_path)
    return out_path
