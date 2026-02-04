from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Optional, Union

import openpyxl


XlsxSource = Union[Path, bytes]


def _load_workbook(source: XlsxSource, *, data_only: bool = True, read_only: bool = False) -> openpyxl.Workbook:
    if isinstance(source, Path):
        return openpyxl.load_workbook(source, data_only=data_only, read_only=read_only)
    if isinstance(source, (bytes, bytearray)):
        return openpyxl.load_workbook(BytesIO(source), data_only=data_only, read_only=read_only)
    raise TypeError("Unsupported xlsx source type")


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_header(text: Any) -> str:
    s = _to_str(text)
    if not s:
        return ""
    # Normalize common variations: remove spaces, unify punctuation/case.
    s = (
        s.replace("\u3000", " ")
        .replace(" ", "")
        .replace("：", ":")
        .replace("／", "/")
        .replace("-", "")
        .replace("_", "")
    )
    return s.lower()


def _build_index(header: list[Any]) -> dict[str, int]:
    indices: dict[str, int] = {}
    for i, name in enumerate(header, start=1):
        key = _norm_header(name)
        if not key or key in indices:
            continue
        indices[key] = i
    return indices


def _find_first(indices: dict[str, int], candidates: Iterable[str]) -> Optional[int]:
    for c in candidates:
        k = _norm_header(c)
        if k in indices:
            return indices[k]
    return None


def _parse_datetime(value: Any) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    text = _to_str(value)
    if not text:
        return None
    # Hotcake exports: "2026/01/01 10:00" or "2025/12/30 18:49"
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def _parse_date(value: Any) -> Optional[datetime]:
    dt = _parse_datetime(value)
    if dt is not None:
        return dt.replace(hour=0, minute=0, second=0, microsecond=0)
    text = _to_str(value)
    if not text:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            d = datetime.strptime(text, fmt)
            return d.replace(hour=0, minute=0, second=0, microsecond=0)
        except ValueError:
            pass
    return None


def _to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        text = _to_str(value).replace(",", "").replace("元", "")
        if not text:
            return 0.0
        return float(text)


@dataclass(frozen=True)
class OrdersRow:
    order_id: str
    order_code: str
    service_start: datetime
    store: str
    designer: str
    service: str
    order_status: str
    checkin_time: Optional[datetime]
    bill_id: str
    bill_amount: float
    member_name: str
    phone: str


def _load_hotcake_orders_with_columns(ws, columns: dict[str, int], order_code_col: Optional[int]) -> list[OrdersRow]:
    rows: list[OrdersRow] = []
    for r in range(2, ws.max_row + 1):
        order_id = _to_str(ws.cell(r, columns["order_id"]).value)
        if not order_id:
            continue
        service_start = _parse_datetime(ws.cell(r, columns["service_start"]).value)
        if service_start is None:
            continue
        store = _to_str(ws.cell(r, columns["store"]).value)
        designer = _to_str(ws.cell(r, columns["designer"]).value)
        service = _to_str(ws.cell(r, columns["service"]).value)
        order_status = _to_str(ws.cell(r, columns["status"]).value)
        checkin_time = _parse_datetime(ws.cell(r, columns["checkin_time"]).value)
        member_name = _to_str(ws.cell(r, columns["member_name"]).value)
        phone = _to_str(ws.cell(r, columns["phone"]).value)
        bill_id = _to_str(ws.cell(r, columns["bill_id"]).value)
        bill_amount = _to_float(ws.cell(r, columns["bill_amount"]).value)
        order_code = ""
        if order_code_col is not None:
            order_code = _to_str(ws.cell(r, order_code_col).value)

        rows.append(
            OrdersRow(
                order_id=order_id,
                order_code=order_code,
                service_start=service_start,
                store=store,
                designer=designer,
                service=service,
                order_status=order_status,
                checkin_time=checkin_time,
                bill_id=bill_id,
                bill_amount=bill_amount,
                member_name=member_name,
                phone=phone,
            )
        )
    return rows


def load_hotcake_orders_xlsx_with_mapping(
    source: XlsxSource,
    *,
    sheet_name: Optional[str],
    mapping: dict[str, int],
    order_code_col: Optional[int] = None,
) -> list[OrdersRow]:
    wb = _load_workbook(source, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    return _load_hotcake_orders_with_columns(ws, mapping, order_code_col)


def load_hotcake_orders_xlsx(source: XlsxSource) -> list[OrdersRow]:
    wb = _load_workbook(source, data_only=True)
    if "訂單報表" in wb.sheetnames:
        ws = wb["訂單報表"]
    else:
        ws = wb[wb.sheetnames[0]]

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    indices = _build_index(header)

    order_id_col = _find_first(indices, ["訂單編號"])
    if order_id_col is None:
        raise ValueError("Hotcake 訂單報表缺少欄位：訂單編號")

    service_start_col = _find_first(indices, ["日期時間", "服務日期時間", "服務開始時間", "開始時間"])
    store_col = _find_first(indices, ["分店", "門市", "店別"])
    designer_col = _find_first(indices, ["設計師", "師傅", "服務人員"])
    service_col = _find_first(indices, ["服務", "服務項目", "項目"])
    status_col = _find_first(indices, ["訂單狀態", "狀態"])
    checkin_col = _find_first(indices, ["報到/取消時間", "報到取消時間", "報到/取消時間", "報到/取消時間"])
    member_name_col = _find_first(indices, ["會員姓名", "姓名"])
    phone_col = _find_first(indices, ["手機號碼", "電話號碼", "手機", "電話"])
    bill_id_col = _find_first(indices, ["帳單編號"])
    bill_amount_col = _find_first(indices, ["帳單金額", "結帳金額", "帳單總額"])

    missing_cols = [
        ("日期時間", service_start_col),
        ("分店", store_col),
        ("設計師", designer_col),
        ("服務", service_col),
        ("訂單狀態", status_col),
        ("報到/取消 時間", checkin_col),
        ("會員姓名", member_name_col),
        ("手機號碼", phone_col),
        ("帳單編號", bill_id_col),
        ("帳單金額", bill_amount_col),
    ]
    missing = [name for name, col in missing_cols if col is None]
    if missing:
        raise ValueError(f"Hotcake 訂單報表缺少欄位: {missing}")

    # Second 訂單編號 column (code) may exist.
    order_code_col: Optional[int] = None
    seen = 0
    for i, name in enumerate(header, start=1):
        if _norm_header(name) == _norm_header("訂單編號"):
            seen += 1
            if seen == 2:
                order_code_col = i
                break

    columns = {
        "order_id": order_id_col,
        "service_start": service_start_col,
        "store": store_col,
        "designer": designer_col,
        "service": service_col,
        "status": status_col,
        "checkin_time": checkin_col,
        "member_name": member_name_col,
        "phone": phone_col,
        "bill_id": bill_id_col,
        "bill_amount": bill_amount_col,
    }
    return _load_hotcake_orders_with_columns(ws, columns, order_code_col)


@dataclass(frozen=True)
class BillRow:
    bill_id: str
    settlement_time: datetime
    attributed_date: datetime
    store: str
    designer: str
    item: str
    cash: float
    credit_card: float
    linepay: float
    bill_amount: float


def _load_hotcake_bill_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[BillRow]:
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    indices = _build_index(header)

    bill_id_col = _find_first(indices, ["帳單編號"])
    settle_col = _find_first(indices, ["結帳操作時間", "結帳時間", "操作時間"])
    attr_col = _find_first(indices, ["計算歸屬日", "歸屬日"])
    store_col = _find_first(indices, ["分店", "門市", "店別"])
    designer_col = _find_first(indices, ["設計師", "師傅", "服務人員"])
    item_col = _find_first(indices, ["項目", "服務項目", "商品名稱"])
    cash_col = _find_first(indices, ["現金", "現金支付", "現金收款"])
    credit_col = _find_first(indices, ["信用卡", "刷卡"])
    linepay_col = _find_first(indices, ["Linepay", "Line Pay", "LinePay", "LINEPAY", "LINE PAY", "linepay", "line pay"])
    amount_col = _find_first(indices, ["結帳金額", "帳單金額", "應收金額"])

    missing_cols = [
        ("帳單編號", bill_id_col),
        ("結帳操作時間", settle_col),
        ("計算歸屬日", attr_col),
        ("分店", store_col),
        ("設計師", designer_col),
        ("項目", item_col),
        ("現金", cash_col),
        ("結帳金額", amount_col),
    ]
    missing = [name for name, col in missing_cols if col is None]
    if missing:
        raise ValueError(f"Hotcake 帳單紀錄缺少欄位: {missing}")

    rows: list[BillRow] = []
    for r in range(2, ws.max_row + 1):
        bill_id = _to_str(ws.cell(r, bill_id_col).value)
        if not bill_id:
            continue
        settlement_time = _parse_datetime(ws.cell(r, settle_col).value)
        attributed_date = _parse_date(ws.cell(r, attr_col).value)
        if settlement_time is None or attributed_date is None:
            continue
        store = _to_str(ws.cell(r, store_col).value)
        designer = _to_str(ws.cell(r, designer_col).value)
        item = _to_str(ws.cell(r, item_col).value)
        cash = _to_float(ws.cell(r, cash_col).value)
        credit_card = _to_float(ws.cell(r, credit_col).value) if credit_col else 0.0
        linepay = _to_float(ws.cell(r, linepay_col).value) if linepay_col else 0.0
        bill_amount = _to_float(ws.cell(r, amount_col).value)
        rows.append(
            BillRow(
                bill_id=bill_id,
                settlement_time=settlement_time,
                attributed_date=attributed_date,
                store=store,
                designer=designer,
                item=item,
                cash=cash,
                credit_card=credit_card,
                linepay=linepay,
                bill_amount=bill_amount,
            )
        )
    return rows


@dataclass(frozen=True)
class HotcakeBills:
    service: list[BillRow]
    topup: list[BillRow]


def load_hotcake_bills_xlsx_with_mapping(
    source: XlsxSource,
    *,
    service_sheet: str,
    topup_sheet: str,
    mapping: dict[str, int],
) -> HotcakeBills:
    wb = _load_workbook(source, data_only=True)
    if service_sheet not in wb.sheetnames:
        raise ValueError(f"找不到分頁：{service_sheet}")
    if topup_sheet not in wb.sheetnames:
        raise ValueError(f"找不到分頁：{topup_sheet}")

    def _load_with_mapping(ws):
        rows: list[BillRow] = []
        for r in range(2, ws.max_row + 1):
            bill_id = _to_str(ws.cell(r, mapping["bill_id"]).value)
            if not bill_id:
                continue
            settlement_time = _parse_datetime(ws.cell(r, mapping["settlement_time"]).value)
            attributed_date = _parse_date(ws.cell(r, mapping["attributed_date"]).value)
            if settlement_time is None or attributed_date is None:
                continue
            store = _to_str(ws.cell(r, mapping["store"]).value)
            designer = _to_str(ws.cell(r, mapping["designer"]).value)
            item = _to_str(ws.cell(r, mapping["item"]).value)
            cash = _to_float(ws.cell(r, mapping["cash"]).value)
            credit_card = _to_float(ws.cell(r, mapping.get("credit_card", 0)).value) if mapping.get("credit_card") else 0.0
            linepay = _to_float(ws.cell(r, mapping.get("linepay", 0)).value) if mapping.get("linepay") else 0.0
            bill_amount = _to_float(ws.cell(r, mapping["bill_amount"]).value)
            rows.append(
                BillRow(
                    bill_id=bill_id,
                    settlement_time=settlement_time,
                    attributed_date=attributed_date,
                    store=store,
                    designer=designer,
                    item=item,
                    cash=cash,
                    credit_card=credit_card,
                    linepay=linepay,
                    bill_amount=bill_amount,
                )
            )
        return rows

    service_rows = _load_with_mapping(wb[service_sheet])
    topup_rows = _load_with_mapping(wb[topup_sheet])
    return HotcakeBills(service=service_rows, topup=topup_rows)


def load_hotcake_bills_xlsx(source: XlsxSource) -> HotcakeBills:
    wb = _load_workbook(source, data_only=True)
    if "服務" not in wb.sheetnames:
        raise ValueError("Hotcake 帳單紀錄找不到「服務」分頁")
    if "儲值金" not in wb.sheetnames:
        raise ValueError("Hotcake 帳單紀錄找不到「儲值金」分頁")

    service_rows = _load_hotcake_bill_sheet(wb["服務"])
    topup_rows = _load_hotcake_bill_sheet(wb["儲值金"])

    return HotcakeBills(service=service_rows, topup=topup_rows)


@dataclass(frozen=True)
class PosRow:
    product_name: str
    created_time: datetime
    terminal_name: str
    order_amount: float
    cash_paid: float
    pay_status: str
    order_status: str
    pay_method: str


def load_pos_history_orders_xlsx_with_mapping(
    source: XlsxSource,
    *,
    sheet_name: Optional[str],
    header_row: int,
    mapping: dict[str, int],
) -> list[PosRow]:
    wb = _load_workbook(source, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows: list[PosRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        product_name = _to_str(ws.cell(r, mapping["product_name"]).value)
        if not product_name:
            continue
        created_time = _parse_datetime(ws.cell(r, mapping["created_time"]).value)
        if created_time is None:
            continue
        terminal_name = _to_str(ws.cell(r, mapping["terminal_name"]).value)
        order_amount = _to_float(ws.cell(r, mapping["order_amount"]).value)
        cash_paid = _to_float(ws.cell(r, mapping["cash_paid"]).value)
        pay_status = _to_str(ws.cell(r, mapping["pay_status"]).value)
        order_status = _to_str(ws.cell(r, mapping["order_status"]).value)
        pay_method = _to_str(ws.cell(r, mapping["pay_method"]).value)
        rows.append(
            PosRow(
                product_name=product_name,
                created_time=created_time,
                terminal_name=terminal_name,
                order_amount=order_amount,
                cash_paid=cash_paid,
                pay_status=pay_status,
                order_status=order_status,
                pay_method=pay_method,
            )
        )
    return rows


def load_pos_history_orders_xlsx(source: XlsxSource) -> list[PosRow]:
    wb = _load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Header is at row 3 in current exports, with summaries above.
    header_row = 3
    header = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    indices = _build_index(header)

    product_col = _find_first(indices, ["商品名稱", "品項", "項目"])
    created_col = _find_first(indices, ["建立時間", "建立日期時間", "時間"])
    terminal_col = _find_first(indices, ["機台名稱", "門市", "店別"])
    amount_col = _find_first(indices, ["訂單金額", "應收金額", "金額"])
    cash_col = _find_first(indices, ["現金支付", "現金"])
    pay_status_col = _find_first(indices, ["付款狀態", "支付狀態"])
    order_status_col = _find_first(indices, ["訂單狀態", "狀態"])
    pay_method_col = _find_first(indices, ["付款方式", "支付方式"])

    missing_cols = [
        ("商品名稱", product_col),
        ("建立時間", created_col),
        ("機台名稱", terminal_col),
        ("訂單金額", amount_col),
        ("現金支付", cash_col),
        ("付款狀態", pay_status_col),
        ("訂單狀態", order_status_col),
        ("付款方式", pay_method_col),
    ]
    missing = [name for name, col in missing_cols if col is None]
    if missing:
        raise ValueError(f"收銀機歷史訂單缺少欄位: {missing}")

    rows: list[PosRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        product_name = _to_str(ws.cell(r, product_col).value)
        if not product_name:
            continue
        created_time = _parse_datetime(ws.cell(r, created_col).value)
        if created_time is None:
            continue
        terminal_name = _to_str(ws.cell(r, terminal_col).value)
        order_amount = _to_float(ws.cell(r, amount_col).value)
        cash_paid = _to_float(ws.cell(r, cash_col).value)
        pay_status = _to_str(ws.cell(r, pay_status_col).value)
        order_status = _to_str(ws.cell(r, order_status_col).value)
        pay_method = _to_str(ws.cell(r, pay_method_col).value)
        rows.append(
            PosRow(
                product_name=product_name,
                created_time=created_time,
                terminal_name=terminal_name,
                order_amount=order_amount,
                cash_paid=cash_paid,
                pay_status=pay_status,
                order_status=order_status,
                pay_method=pay_method,
            )
        )
    return rows


@dataclass(frozen=True)
class CardMachineRow:
    order_id: str
    store: str
    device_name: str
    amount: float
    paid_amount: float
    transaction_time: datetime
    pay_method: str


def load_card_machine_xlsx(source: XlsxSource) -> list[CardMachineRow]:
    wb = _load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Header is at row 2 in current exports
    header_row = 2
    header = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    indices = _build_index(header)

    order_col = _find_first(indices, ["訂單編號"])
    store_col = _find_first(indices, ["店鋪名稱", "分店", "門市", "店別"])
    device_col = _find_first(indices, ["設備名稱", "機台名稱"])
    amount_col = _find_first(indices, ["交易金額", "金額"])
    paid_col = _find_first(indices, ["實付金額"])
    time_col = _find_first(indices, ["交易時間", "時間"])
    method_col = _find_first(indices, ["支付方式", "付款方式"])

    missing_cols = [
        ("訂單編號", order_col),
        ("店鋪名稱", store_col),
        ("設備名稱", device_col),
        ("交易金額", amount_col),
        ("實付金額", paid_col),
        ("交易時間", time_col),
        ("支付方式", method_col),
    ]
    missing = [name for name, col in missing_cols if col is None]
    if missing:
        raise ValueError(f"智慧刷卡機交易紀錄缺少欄位: {missing}")

    rows: list[CardMachineRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        order_id = _to_str(ws.cell(r, order_col).value)
        if not order_id:
            continue
        store = _to_str(ws.cell(r, store_col).value)
        device_name = _to_str(ws.cell(r, device_col).value)
        amount = _to_float(ws.cell(r, amount_col).value)
        paid_amount = _to_float(ws.cell(r, paid_col).value)
        transaction_time = _parse_datetime(ws.cell(r, time_col).value)
        if transaction_time is None:
            continue
        pay_method = _to_str(ws.cell(r, method_col).value)
        rows.append(
            CardMachineRow(
                order_id=order_id,
                store=store,
                device_name=device_name,
                amount=amount,
                paid_amount=paid_amount,
                transaction_time=transaction_time,
                pay_method=pay_method,
            )
        )
    return rows
