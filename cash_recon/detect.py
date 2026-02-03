from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Literal, Optional, Union

import openpyxl

from .parse import _build_index, _find_first, _norm_header


Kind = Literal["hotcake_bills", "hotcake_orders", "pos_orders", "unknown"]
XlsxSource = Union[Path, bytes]


@dataclass(frozen=True)
class Detection:
    kind: Kind
    reason: str
    sheet: Optional[str] = None


def detect_xlsx_kind(source: XlsxSource) -> Detection:
    try:
        if isinstance(source, Path):
            wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
        else:
            wb = openpyxl.load_workbook(BytesIO(source), data_only=True, read_only=True)
    except Exception as e:
        return Detection(kind="unknown", reason=f"無法讀取 xlsx: {e}")

    sheetnames = set(wb.sheetnames)

    # Hotcake bills: has sheets 服務 and 儲值金 and a header with 帳單編號
    if "服務" in sheetnames and "儲值金" in sheetnames:
        ws = wb["服務"]
        header = [ws.cell(1, c).value for c in range(1, min(ws.max_column, 80) + 1)]
        idx = _build_index(header)
        if _find_first(idx, ["帳單編號"]) is not None and _find_first(idx, ["結帳操作時間", "結帳時間"]) is not None:
            return Detection(kind="hotcake_bills", reason="包含「服務/儲值金」分頁且表頭符合帳單紀錄", sheet="服務")

    # Hotcake orders: sheet 訂單報表 or first sheet with 訂單編號 + 日期時間 + 訂單狀態
    if "訂單報表" in sheetnames:
        ws = wb["訂單報表"]
        header = [ws.cell(1, c).value for c in range(1, min(ws.max_column, 120) + 1)]
        idx = _build_index(header)
        if _find_first(idx, ["訂單編號"]) is not None and _find_first(idx, ["日期時間", "服務日期時間", "服務開始時間", "開始時間"]) is not None:
            return Detection(kind="hotcake_orders", reason="包含「訂單報表」分頁且表頭符合訂單/預約報表", sheet="訂單報表")

    # POS orders: typically first row contains "歷史訂單", header row 3 contains 商品名稱/建立時間/現金支付
    ws = wb[wb.sheetnames[0]]
    first_cell = ws.cell(1, 1).value
    if isinstance(first_cell, str) and "歷史訂單" in first_cell:
        header_row = 3
        header = [ws.cell(header_row, c).value for c in range(1, min(ws.max_column, 120) + 1)]
        idx = _build_index(header)
        if _find_first(idx, ["商品名稱"]) is not None and _find_first(idx, ["建立時間"]) is not None and _find_first(idx, ["現金支付"]) is not None:
            return Detection(kind="pos_orders", reason="表頭/格式符合收銀機歷史訂單", sheet=wb.sheetnames[0])

    # Fallback: try classify by presence of key headers on first sheet row 1 / row 3
    ws = wb[wb.sheetnames[0]]
    header1 = [ws.cell(1, c).value for c in range(1, min(ws.max_column, 120) + 1)]
    idx1 = _build_index(header1)
    if _find_first(idx1, ["訂單編號"]) is not None and _find_first(idx1, ["日期時間", "服務開始時間"]) is not None:
        return Detection(kind="hotcake_orders", reason="首分頁表頭符合 Hotcake 訂單/預約報表", sheet=wb.sheetnames[0])

    header3 = [ws.cell(3, c).value for c in range(1, min(ws.max_column, 120) + 1)]
    idx3 = _build_index(header3)
    if _find_first(idx3, ["商品名稱"]) is not None and _find_first(idx3, ["建立時間"]) is not None and _find_first(idx3, ["現金支付"]) is not None:
        return Detection(kind="pos_orders", reason="首分頁第 3 列表頭符合收銀機歷史訂單", sheet=wb.sheetnames[0])

    return Detection(kind="unknown", reason="無法判斷報表類型（可能是格式更新或非支援報表）")
