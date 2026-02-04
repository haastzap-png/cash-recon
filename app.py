from __future__ import annotations

from datetime import datetime
from io import BytesIO
import os
from typing import Optional

import openpyxl
import streamlit as st

from cash_recon.detect import detect_xlsx_kind
from cash_recon.logic import build_cash_recon
from cash_recon.models import Period
from cash_recon.parse import (
    load_hotcake_bills_xlsx,
    load_hotcake_bills_xlsx_with_mapping,
    load_hotcake_orders_xlsx,
    load_hotcake_orders_xlsx_with_mapping,
    load_pos_history_orders_xlsx,
    load_pos_history_orders_xlsx_with_mapping,
)
from cash_recon.report import build_cash_recon_workbook


def _parse_dt(text: str) -> Optional[datetime]:
    text = (text or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def _require_passcode():
    passcode = None
    try:
        passcode = st.secrets.get("passcode")
    except Exception:
        pass
    passcode = passcode or os.environ.get("APP_PASSCODE")

    if not passcode:
        st.error("尚未設定通關碼。請在 Streamlit Secrets 設定 `passcode`，或在環境變數設定 `APP_PASSCODE`。")
        st.stop()

    code = st.text_input("通關碼", type="password")
    if not code:
        st.stop()
    if code != passcode:
        st.error("通關碼不正確")
        st.stop()


def _get_headers(xlsx_bytes: bytes, header_row: int = 1, sheet_name: Optional[str] = None) -> tuple[list[str], str]:
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    sheet = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    max_col = min(ws.max_column, 200)
    headers = [ws.cell(header_row, c).value for c in range(1, max_col + 1)]
    headers = [("" if h is None else str(h)) for h in headers]
    return headers, sheet


def _header_options(headers: list[str]) -> list[tuple[str, int]]:
    options: list[tuple[str, int]] = []
    for i, h in enumerate(headers, start=1):
        label = f"第{i}欄 - {h}" if h else f"第{i}欄 - (空白)"
        options.append((label, i))
    return options


def _default_index(headers: list[str], candidates: list[str]) -> int:
    norm = lambda s: s.replace(" ", "").replace("\u3000", "").lower()
    header_norm = [norm(h) for h in headers]
    for c in candidates:
        cn = norm(c)
        if cn in header_norm:
            return header_norm.index(cn)
    return 0


def _reset_app():
    st.session_state["upload_key"] = st.session_state.get("upload_key", 0) + 1
    st.session_state.pop("result_ready", None)
    st.rerun()


st.set_page_config(page_title="現金對帳表", layout="wide")
st.title("現金對帳表 (MVP)")

_require_passcode()

st.markdown(
    """
上傳 Hotcake 的 **帳單紀錄** + **訂單(預約)報表**，可選上傳 **收銀機歷史訂單**。
輸出重點：該區間每間店 **要收回來多少現金**，以及 **漏結帳清單** 是否為空。
"""
)

with st.sidebar:
    st.header("輸入")
    store_options = ["中壢三光店", "桃園中平店", "楊梅四維店", "中壢體育園區店", "其他(手動輸入)"]
    store_choice = st.selectbox("分店", options=store_options, index=0)
    if store_choice == "其他(手動輸入)":
        store = st.text_input("分店名稱（手動輸入）", value="")
    else:
        store = store_choice
    start_text = st.text_input("區間開始 (YYYY-MM-DD HH:MM)", value="2026-01-01 00:00")
    end_text = st.text_input("區間結束 (YYYY-MM-DD HH:MM)", value="2026-01-31 23:59")
    topup_mode = "settlement_time"
    st.caption("儲值金：依結帳操作時間計入")
    time_tolerance = st.number_input("POS 對帳時間容忍(分鐘)", min_value=0, max_value=240, value=120)
    auto_clear = st.checkbox("下載後自動清空", value=True)
    if st.button("清空/重置"):
        _reset_app()

st.header("上傳檔案")
if "upload_key" not in st.session_state:
    st.session_state["upload_key"] = 0

uploads = st.file_uploader(
    "把 xlsx 拖進來（可一次上傳多個，系統會自動判斷：帳單紀錄/訂單(預約)/收銀機歷史訂單）",
    type=["xlsx"],
    accept_multiple_files=True,
    key=f"uploads_{st.session_state['upload_key']}",
)

hotcake_bills_up = None
hotcake_orders_up = None
pos_orders_up = None

if uploads:
    st.subheader("辨識結果")
    assignments = []
    for up in uploads:
        content = up.getvalue()
        d = detect_xlsx_kind(content)
        choice = st.selectbox(
            f"{up.name}（自動判斷：{d.kind}）",
            options=[
                ("自動判斷", "auto"),
                ("Hotcake 帳單紀錄", "hotcake_bills"),
                ("Hotcake 訂單/預約報表", "hotcake_orders"),
                ("收銀機 歷史訂單", "pos_orders"),
                ("忽略此檔", "ignore"),
            ],
            format_func=lambda x: x[0],
            key=f"type_{up.name}",
        )
        assignments.append((up, d.kind, choice[1]))

    for up, detected, manual in assignments:
        kind = detected if manual == "auto" else manual
        if kind == "hotcake_bills" and hotcake_bills_up is None:
            hotcake_bills_up = up
        elif kind == "hotcake_orders" and hotcake_orders_up is None:
            hotcake_orders_up = up
        elif kind == "pos_orders" and pos_orders_up is None:
            pos_orders_up = up

    st.subheader("進階設定（欄位變動時使用）")
    with st.expander("手動欄位對應"):
        manual_orders = st.checkbox("手動欄位對應：Hotcake 訂單/預約報表", value=False)
        manual_bills = st.checkbox("手動欄位對應：Hotcake 帳單紀錄", value=False)
        manual_pos = st.checkbox("手動欄位對應：收銀機 歷史訂單", value=False) if pos_orders_up else False

        if manual_orders and hotcake_orders_up:
            headers, sheet_name = _get_headers(hotcake_orders_up.getvalue(), header_row=1, sheet_name="訂單報表")
            opts = _header_options(headers)
            def _pick(label, candidates):
                idx = _default_index(headers, candidates)
                return st.selectbox(label, options=opts, format_func=lambda x: x[0], index=idx, key=f"o_{label}")[1]
            st.session_state["orders_mapping"] = {
                "order_id": _pick("訂單編號", ["訂單編號"]),
                "service_start": _pick("日期時間/服務開始", ["日期時間", "服務開始時間", "開始時間"]),
                "store": _pick("分店", ["分店", "門市", "店別"]),
                "designer": _pick("設計師/師傅", ["設計師", "師傅", "服務人員"]),
                "service": _pick("服務/項目", ["服務", "服務項目", "項目"]),
                "status": _pick("訂單狀態", ["訂單狀態", "狀態"]),
                "checkin_time": _pick("報到/取消時間", ["報到/取消時間", "報到取消時間"]),
                "member_name": _pick("會員姓名", ["會員姓名", "姓名"]),
                "phone": _pick("手機號碼", ["手機號碼", "電話號碼", "手機", "電話"]),
                "bill_id": _pick("帳單編號", ["帳單編號"]),
                "bill_amount": _pick("帳單金額", ["帳單金額", "結帳金額"]),
            }
            st.session_state["orders_sheet"] = sheet_name
        if not manual_orders:
            st.session_state.pop("orders_mapping", None)
            st.session_state.pop("orders_sheet", None)

        if manual_bills and hotcake_bills_up:
            headers, _ = _get_headers(hotcake_bills_up.getvalue(), header_row=1, sheet_name="服務")
            opts = _header_options(headers)
            def _pick_b(label, candidates):
                idx = _default_index(headers, candidates)
                return st.selectbox(label, options=opts, format_func=lambda x: x[0], index=idx, key=f"b_{label}")[1]
            st.session_state["bills_mapping"] = {
                "bill_id": _pick_b("帳單編號", ["帳單編號"]),
                "settlement_time": _pick_b("結帳操作時間", ["結帳操作時間", "結帳時間", "操作時間"]),
                "attributed_date": _pick_b("計算歸屬日", ["計算歸屬日", "歸屬日"]),
                "store": _pick_b("分店", ["分店", "門市", "店別"]),
                "designer": _pick_b("設計師/師傅", ["設計師", "師傅", "服務人員"]),
                "item": _pick_b("項目", ["項目", "服務項目", "商品名稱"]),
                "cash": _pick_b("現金", ["現金", "現金支付", "現金收款"]),
                "bill_amount": _pick_b("結帳金額", ["結帳金額", "帳單金額"]),
            }
        if not manual_bills:
            st.session_state.pop("bills_mapping", None)

        if manual_pos and pos_orders_up:
            header_row = st.number_input("收銀機表頭列(預設 3)", min_value=1, max_value=10, value=3)
            headers, sheet_name = _get_headers(pos_orders_up.getvalue(), header_row=header_row, sheet_name=None)
            opts = _header_options(headers)
            def _pick_p(label, candidates):
                idx = _default_index(headers, candidates)
                return st.selectbox(label, options=opts, format_func=lambda x: x[0], index=idx, key=f"p_{label}")[1]
            st.session_state["pos_mapping"] = {
                "product_name": _pick_p("商品名稱", ["商品名稱", "品項", "項目"]),
                "created_time": _pick_p("建立時間", ["建立時間", "建立日期時間", "時間"]),
                "terminal_name": _pick_p("機台名稱", ["機台名稱", "門市", "店別"]),
                "order_amount": _pick_p("訂單金額", ["訂單金額", "應收金額", "金額"]),
                "cash_paid": _pick_p("現金支付", ["現金支付", "現金"]),
                "pay_status": _pick_p("付款狀態", ["付款狀態", "支付狀態"]),
                "order_status": _pick_p("訂單狀態", ["訂單狀態", "狀態"]),
                "pay_method": _pick_p("付款方式", ["付款方式", "支付方式"]),
            }
            st.session_state["pos_sheet"] = sheet_name
            st.session_state["pos_header_row"] = header_row
        if not manual_pos:
            st.session_state.pop("pos_mapping", None)
            st.session_state.pop("pos_sheet", None)
            st.session_state.pop("pos_header_row", None)

if not uploads:
    st.info("請先上傳檔案。至少需要 Hotcake：帳單紀錄 + 訂單/預約報表。")

run = st.button("產出現金對帳表", type="primary", disabled=not (hotcake_bills_up and hotcake_orders_up))

if run:
    start_dt = _parse_dt(start_text)
    end_dt = _parse_dt(end_text)
    if start_dt is None or end_dt is None or start_dt > end_dt:
        st.error("區間日期時間格式不正確，請用 YYYY-MM-DD HH:MM 或 YYYY-MM-DD HH:MM:SS，且開始需早於結束。")
        st.stop()

    if not store.strip():
        st.error("請輸入分店名稱。")
        st.stop()

    with st.spinner("讀取報表中..."):
        bills_bytes = hotcake_bills_up.getvalue()
        orders_bytes = hotcake_orders_up.getvalue()
        pos_bytes = pos_orders_up.getvalue() if pos_orders_up else None

        try:
            if st.session_state.get("orders_mapping"):
                orders = load_hotcake_orders_xlsx_with_mapping(
                    orders_bytes,
                    sheet_name=st.session_state.get("orders_sheet"),
                    mapping=st.session_state["orders_mapping"],
                )
            else:
                orders = load_hotcake_orders_xlsx(orders_bytes)

            if st.session_state.get("bills_mapping"):
                bills = load_hotcake_bills_xlsx_with_mapping(
                    bills_bytes,
                    service_sheet="服務",
                    topup_sheet="儲值金",
                    mapping=st.session_state["bills_mapping"],
                )
            else:
                bills = load_hotcake_bills_xlsx(bills_bytes)

            if pos_bytes:
                if st.session_state.get("pos_mapping"):
                    pos_orders = load_pos_history_orders_xlsx_with_mapping(
                        pos_bytes,
                        sheet_name=st.session_state.get("pos_sheet"),
                        header_row=st.session_state.get("pos_header_row", 3),
                        mapping=st.session_state["pos_mapping"],
                    )
                else:
                    pos_orders = load_pos_history_orders_xlsx(pos_bytes)
            else:
                pos_orders = None
        except Exception as e:
            st.error(f"讀取報表失敗：{e}")
            st.info("若表頭有變動，請展開「進階設定」→「手動欄位對應」後再試。")
            st.stop()

    with st.spinner("計算中..."):
        result = build_cash_recon(
            period=Period(start=start_dt, end=end_dt),
            store=store.strip(),
            orders=orders,
            bills=bills,
            pos_orders=pos_orders,
            topup_mode=topup_mode,
            time_tolerance_minutes=int(time_tolerance),
        )

    st.header("結果")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Hotcake 服務現金", f"{result.totals.hotcake_service_cash:,.0f}")
    c2.metric("Hotcake 儲值金現金", f"{result.totals.hotcake_topup_cash:,.0f}")
    c3.metric("Hotcake 現金合計", f"{result.totals.hotcake_cash_total:,.0f}")
    if result.totals.pos_cash_total is not None:
        c4.metric("收銀機現金合計", f"{result.totals.pos_cash_total:,.0f}")
    else:
        c4.metric("收銀機現金合計", "（未上傳）")

    if result.missing_bills:
        st.error(f"漏結帳清單不為空：{len(result.missing_bills)} 筆。依你的規則，現金對帳表不可視為正確。")
        st.dataframe(
            [
                {
                    "分店": r.store,
                    "日期": r.service_start.strftime("%Y-%m-%d"),
                    "日期時間(服務開始)": r.service_start.strftime("%Y-%m-%d %H:%M:%S"),
                    "師傅": r.designer,
                    "服務": r.service,
                    "訂單編號": r.order_id,
                    "訂單代碼": r.order_code,
                    "會員姓名": r.member_name,
                    "手機號碼": r.phone,
                }
                for r in result.missing_bills
            ],
            use_container_width=True,
        )
    else:
        st.success("漏結帳清單為空：可視為正確現金對帳表。")

    if result.totals.pos_cash_diff is not None:
        st.info(f"收銀機現金 - Hotcake 現金合計：{result.totals.pos_cash_diff:,.0f}")

    if pos_orders is not None:
        st.subheader("時間容忍外的資料")
        c5, c6 = st.columns(2)
        c5.metric("Hotcake 未匹配筆數", f"{len(result.hotcake_time_mismatches)}")
        c6.metric("POS 未匹配筆數", f"{len(result.pos_time_mismatches)}")
        if result.hotcake_time_mismatches:
            st.dataframe(
                [
                    {
                        "分店": r.store,
                        "日期": r.service_start.strftime("%Y-%m-%d"),
                        "日期時間": r.service_start.strftime("%Y-%m-%d %H:%M:%S"),
                        "師傅": r.designer,
                        "服務": r.service,
                        "分鐘": r.minutes,
                        "帳單編號": r.bill_id,
                        "帳單金額": r.bill_amount,
                        "現金": r.cash,
                        "最近POS時間": r.nearest_pos_time.strftime("%Y-%m-%d %H:%M:%S") if r.nearest_pos_time else "",
                        "時間差(分)": r.nearest_diff_minutes,
                    }
                    for r in result.hotcake_time_mismatches
                ],
                use_container_width=True,
            )
        if result.pos_time_mismatches:
            st.dataframe(
                [
                    {
                        "機台名稱": r.terminal_name,
                        "日期時間": r.created_time.strftime("%Y-%m-%d %H:%M:%S"),
                        "師傅": r.designer,
                        "商品名稱": r.product_name,
                        "分鐘": r.minutes,
                        "現金支付": r.cash_paid,
                        "最近Hotcake時間": r.nearest_hotcake_time.strftime("%Y-%m-%d %H:%M:%S") if r.nearest_hotcake_time else "",
                        "時間差(分)": r.nearest_diff_minutes,
                    }
                    for r in result.pos_time_mismatches
                ],
                use_container_width=True,
            )

    wb = build_cash_recon_workbook(result)
    out_name = f"現金對帳表_{store.strip()}_{start_dt.strftime('%Y%m%d')}-{end_dt.strftime('%Y%m%d')}.xlsx"
    bio = BytesIO()
    wb.save(bio)
    report_bytes = bio.getvalue()

    download_kwargs = {}
    if auto_clear:
        download_kwargs["on_click"] = _reset_app
    st.download_button(
        "下載現金對帳表 .xlsx",
        data=report_bytes,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        **download_kwargs,
    )
